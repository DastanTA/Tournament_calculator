import re
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from typing import List, Dict, Tuple


pattern = r'^\d+(\.\d+)?:\d+(\.\d+)?$'


def get_winners(players: List[Dict], is_football=False):
    winners = {}

    sorted_players = sorted(players, key=lambda x: x["score"], reverse=True)

    to_start = 0
    second_parameter = "побед в играх"
    if is_football:
        second_parameter = "количество голов"

    nl_n_start_n = get_places_n_to_start(sorted_players, to_start)
    first_place = nl_n_start_n.get("new_list")
    to_start = nl_n_start_n.get("to_start")
    winners_to_str(first_place, second_parameter, "first place", winners=winners)

    nl_n_start_n = get_places_n_to_start(sorted_players, to_start)
    second_place = nl_n_start_n.get("new_list")
    to_start = nl_n_start_n.get("to_start")
    winners_to_str(second_place, second_parameter, "second place", winners=winners)

    third_place = get_places_n_to_start(sorted_players, to_start).get("new_list")
    winners_to_str(third_place, second_parameter, "third place", winners=winners)

    return winners


def winners_to_str(winners_list, sp: str, place: str, winners: Dict):
    if len(winners_list) <= 1:
        winners[place] = f"{winners_list[0].get('name')}: очков: {winners_list[0].get('score')}, " \
                         f"побед в матчах: {winners_list[0].get('match_wins')}, " \
                         f"ничей: {winners_list[0].get('draw/ничья')}, " \
                         f"{sp}: {winners_list[0].get('game_wins')}, проигрышей: {winners_list[0].get('loose')}"
    elif len(winners_list) > 1:
        winners[place] = []
        for winner in winners_list:
            winners[place].append(
                f"{winner.get('name')}: очков: {winner.get('score')}, побед в матчах: {winner.get('match_wins')}, "
                f"ничей: {winners_list[0].get('draw/ничья')}, "
                f"{sp}: {winner.get('game_wins')}, проигрышей: {winners_list[0].get('loose')}"
            )


def get_places_n_to_start(sorted_list: List[Dict], starting_point) -> Dict:
    list_to_append = [sorted_list[starting_point]]
    to_start = starting_point
    for i in range(starting_point, len(sorted_list)):
        if i != len(sorted_list) - 1:
            if sorted_list[i].get("score") == sorted_list[i + 1].get("score"):
                list_to_append.append(sorted_list[i + 1])
            else:
                to_start = i + 1
                break
    return {"to_start": to_start, "new_list": list_to_append}


def handle_scores(worksheet, contestant, row):
    for result in worksheet[row]:
        if result.value and re.match(pattern, result.value):
            print(f"result.value_len = {len(result.value)}, value = {result.value}")
            result_lst = result.value.split(":")
            first_num = float(result_lst[0])
            second_num = float(result_lst[1])
            contestant["game_wins"] += first_num
            if first_num > second_num:
                contestant["match_wins"] += 1
                contestant["score"] += 3
            elif first_num == second_num:
                contestant["draw/ничья"] += 1
                contestant["score"] += 1
            else:
                contestant["score"] += 0
                contestant["loose"] += 1
    return contestant


def handling_tournament(workbook, tournament_var, tournament_name):
    ws = workbook[tournament_name]
    contestants_number = 0
    contestants = []
    for element in ws["A"]:
        if element.value:
            contestant = {
                "name": element.value,
                "game_wins": 0,
                "match_wins": 0,
                "draw/ничья": 0,
                "loose": 0,
                "score": 0
            }
            contestants_number += 1
            if tournament_name == "Футбол":
                for team in ws["C"]:
                    if team.value:
                        if element.value[-2] == team.value[-1]:
                            row = int(team.coordinate[-1])
                            contestant = handle_scores(worksheet=ws, contestant=contestant, row=row)
            elif tournament_name != "Футбол":
                row = int(element.coordinate[1:])
                contestant = handle_scores(worksheet=ws, contestant=contestant, row=row)
            contestants.append(contestant)
    tournament_var["contestants"] = contestants
    tournament_var["contestants_number"] = contestants_number

    if tournament_name == "Футбол":
        tournament_var["winners"] = get_winners(players=contestants, is_football=True)
    if tournament_name != "Футбол":
        tournament_var["winners"] = get_winners(players=contestants, is_football=False)

    return tournament_var


class Calculator:
    def __init__(self):
        self.root = tk.Tk()
        self.chess = {
            "contestants_number": 0,
            "contestants": [],
            "winners": {}
        }

        self.football = {
            "contestants_number": 0,
            "contestants": [],
            "winners": {}
        }

        self.tennis_a = {
            "contestants_number": 0,
            "contestants": [],
            "winners": {}
        }

        self.tennis_b = {
            "contestants_number": 0,
            "contestants": [],
            "winners": {}
        }

        self.root.geometry("600x600")
        self.root.title("Калькулятор победителей")
        self.label = tk.Label(self.root, text="Загрузи Excel файл!", font=('Arial', 18))
        self.label.pack(padx=10, pady=10)
        self.upload_btn = tk.Button(self.root, text="Загрузить", command=self.open).pack(padx=10, pady=10)

        self.root.mainloop()

    def create_table(self, top, lst_of_tuples: List[Tuple]):
        for i in range(len(lst_of_tuples)):
            for j in range(7):
                e = tk.Entry(top, width=20, fg='black', font=('Arial', 16, 'bold'))
                e.grid(row=i, column=j)
                e.insert(tk.END, lst_of_tuples[i][j])

    def get_list_of_tuples(self, conts: List[Dict], is_football=False) -> List[Tuple]:
        to_return = []
        if is_football:
            first_row = ("#", "Имя", "очки", "побед в матчах", "количество голов", "ничей", "поражений")
        else:
            first_row = ("#", "Имя", "очки", "побед в матчах", "выигранных игр", "ничей", "поражений")
        to_return.append(first_row)
        sorted_conts = sorted(conts, key=lambda x: x["score"], reverse=True)

        for player in sorted_conts:
            tuuple = (sorted_conts.index(player) + 1, player.get("name"), player.get("score"), player.get("match_wins"),
                      player.get("game_wins"), player.get("draw/ничья"), player.get("loose"))
            to_return.append(tuuple)
        return to_return

    def show_chess(self):
        top = tk.Toplevel()
        top.title("Результаты")
        top.geometry("1800x700")
        lst_of_tuples_chess = self.get_list_of_tuples(self.chess.get("contestants"))
        self.create_table(top, lst_of_tuples_chess)

    def show_football(self):
        top = tk.Toplevel()
        top.title("Результаты")
        top.geometry("1800x700")
        lst_of_tuples_chess = self.get_list_of_tuples(self.football.get("contestants"), is_football=True)
        self.create_table(top, lst_of_tuples_chess)

    def show_tennis_a(self):
        top = tk.Toplevel()
        top.title("Результаты")
        top.geometry("1800x700")
        lst_of_tuples_chess = self.get_list_of_tuples(self.tennis_a.get("contestants"))
        self.create_table(top, lst_of_tuples_chess)

    def show_tennis_b(self):
        top = tk.Toplevel()
        top.title("Результаты")
        top.geometry("1800x700")
        lst_of_tuples_chess = self.get_list_of_tuples(self.tennis_b.get("contestants"))
        self.create_table(top, lst_of_tuples_chess)

    def get_xl(self):
        to_download = Workbook()

        sheet_chess = to_download.active
        sheet_chess.title = "Результаты по Шахматам"
        self.enter_data(sheet=sheet_chess, contestants=self.chess.get("contestants"))

        sheet_football = to_download.create_sheet(title="Результаты по Футболу")
        self.enter_data(sheet=sheet_football, contestants=self.football.get("contestants"), is_football=True)

        sheet_tennis_a = to_download.create_sheet(title="Результаты по Теннису (А)")
        self.enter_data(sheet=sheet_tennis_a, contestants=self.tennis_a.get("contestants"))

        sheet_tennis_b = to_download.create_sheet(title="Результаты по Теннису (Б)")
        self.enter_data(sheet=sheet_tennis_b, contestants=self.tennis_b.get("contestants"))

        to_download.save("Результаты_Турнира.xlsx")

        messagebox.showinfo("Успех", "Результаты игр сохранын в Excel файл в той же директории "
                                     "где находится эта программа. Каждый вид игр - в отдельной вкладке")

    def enter_data(self, sheet, contestants, is_football=False):

        sorted_conts = sorted(contestants, key=lambda x: x["score"], reverse=True)

        sheet["A1"] = "#"
        sheet["C1"] = "Очков"
        sheet["D1"] = "Выигранных матчей"
        sheet ["F1"] = "Ничей"
        sheet["G1"] = "Поражений"
        if is_football:
            sheet["B1"] = "Название команды"
            sheet["E1"] = "количество голов"
        else:
            sheet["B1"] = "Имя"
            sheet["E1"] = "Вигранных игр"

        counter = 1
        for player in sorted_conts:
            for k, v in player.items():
                cell = f"A{counter+1}"
                sheet[cell] = counter
                match k:
                    case "name":
                        cell = f"B{counter + 1}"
                        sheet[cell] = v
                    case "score":
                        cell = f"C{counter + 1}"
                        sheet[cell] = v
                    case "match_wins":
                        cell = f"D{counter + 1}"
                        sheet[cell] = v
                    case "game_wins":
                        cell = f"E{counter + 1}"
                        sheet[cell] = v
                    case "draw/ничья":
                        cell = f"F{counter + 1}"
                        sheet[cell] = v
                    case "loose":
                        cell = f"G{counter + 1}"
                        sheet[cell] = v
            counter += 1

    def open(self):
        self.root.filename = filedialog.askopenfilename(initialdir="/home", title="Выбери excel файл, чорт!",
                                                        filetypes=(("Excel файлы", "*.xlsx"), ("все форматы", "*.*")))
        wb = load_workbook(self.root.filename)
        tournaments = wb.sheetnames

        for tournament in tournaments:
            match tournament:
                case "Футбол":
                    self.football = handling_tournament(
                        workbook=wb,
                        tournament_var=self.football,
                        tournament_name=tournament)
                case "Шахматы":
                    self.chess = handling_tournament(
                        workbook=wb,
                        tournament_var=self.chess,
                        tournament_name="Шахматы")
                case "Теннис (А)":
                    self.tennis_a = handling_tournament(
                        workbook=wb,
                        tournament_var=self.tennis_a,
                        tournament_name="Теннис (А)")
                case "Теннис (Б)":
                    self.tennis_b = handling_tournament(
                        workbook=wb,
                        tournament_var=self.tennis_b,
                        tournament_name="Теннис (Б)")

        self.button_chess = tk.Button(self.root, text="Результаты по Шахматам",
                                      command=self.show_chess).pack(padx=10, pady=10)
        self.button_football = tk.Button(self.root, text="Результаты по Футболу",
                                         command=self.show_football).pack(padx=10, pady=10)
        self.button_tennis_a = tk.Button(self.root, text="Результаты по Теннису (А)",
                                         command=self.show_tennis_a).pack(padx=10, pady=10)
        self.button_tennis_b = tk.Button(self.root, text="Результаты по Теннису (Б)",
                                         command=self.show_tennis_b).pack(padx=10, pady=10)
        self.button_get_xl = tk.Button(self.root, text="Выгрузить все в Excel",
                                         command=self.get_xl, bg="yellow").pack(padx=10, pady=10)


Calculator()
